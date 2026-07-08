import sys
import re
from datetime import datetime

from openpyxl import load_workbook

from payroll_clients.base import (
    make_row, check_balance, print_journal_table,
    append_payroll_log, append_digest_log,
    load_config, _qb_confirm, archive_payroll_pdf,
)

# Fixed column positions in Square's "Company Totals Report" xlsx export.
COL_LABEL          = 0   # Date Range / From / To / work-address labels
COL_EARNINGS_LABEL = 1
COL_PAY            = 3
COL_EE_TAX_LABEL   = 4
COL_EE_TAX_AMT     = 5
COL_ER_TAX_LABEL   = 6
COL_ER_TAX_AMT     = 7
COL_DEDUCTIONS_AMT = 9
COL_REIMB_AMT      = 11
COL_EE_BENEFITS_AMT = 13
COL_ER_BENEFITS_AMT = 15
COL_NET_PAY        = 16

# Tax categories bucketed by which agency the payment ultimately goes to.
# IRS gets federal income withholding plus both halves of Social Security/
# Medicare (employee withholding and the employer match are deposited
# together via EFTPS). EDD gets CA income tax + CA disability withholding.
# UI/ETT bundles the employer-only unemployment-related taxes.
EE_IRS_TAXES    = {"EE Fed. Income", "EE Soc. Security", "EE Medicare", "EE Fed. Additional Medicare"}
EE_EDD_TAXES    = {"EE CA State Income", "EE CA State Disability"}
ER_IRS_TAXES    = {"ER Soc. Security", "ER Medicare"}
ER_UI_ETT_TAXES = {"ER Fed. Unemployment", "ER CA State Employment Training", "ER CA State Unemployment"}


def _num(v) -> float:
    return round(float(v), 2) if isinstance(v, (int, float)) else 0.0


def normalize_check_date(date_str: str) -> str:
    """Square prints dates as 'M/D/YY' — normalize to 'MM/DD/YYYY' to match
    the convention every other payroll module's check_date uses."""
    return datetime.strptime(date_str.strip(), "%m/%d/%y").strftime("%m/%d/%Y")


def parse_workbook(xlsx_path: str) -> dict:
    """Parse a Square Payroll 'Company Totals Report' xlsx export.

    The report always leads with a company-wide "All Work Addresses" block
    before breaking the same numbers down per work address — this parser
    only reads that first (aggregate) block, since QuickBooks only needs
    one journal entry per check date regardless of how many locations ran.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, r in enumerate(rows) if r and r[COL_LABEL] == "Date Range"), None
    )
    if header_idx is None:
        raise ValueError(f"{xlsx_path}: not a recognized Square Payroll Company Totals export")

    result = {
        "pay": 0.0, "employer_taxes": 0.0, "net_pay": 0.0,
        "ee_irs": 0.0, "ee_edd": 0.0, "er_irs": 0.0, "er_ui_ett": 0.0,
        "deductions": 0.0, "reimbursements": 0.0, "ee_benefits": 0.0, "er_benefits": 0.0,
    }
    from_date = to_date = None

    for row in rows[header_idx + 1:]:
        label = row[COL_LABEL].strip() if isinstance(row[COL_LABEL], str) else ""
        m = re.match(r"From:\s*(\S+)", label)
        if m: from_date = m.group(1)
        m = re.match(r"To:\s*(\S+)", label)
        if m: to_date = m.group(1)

        if row[COL_EE_TAX_LABEL] in EE_IRS_TAXES:
            result["ee_irs"] += _num(row[COL_EE_TAX_AMT])
        elif row[COL_EE_TAX_LABEL] in EE_EDD_TAXES:
            result["ee_edd"] += _num(row[COL_EE_TAX_AMT])

        if row[COL_ER_TAX_LABEL] in ER_IRS_TAXES:
            result["er_irs"] += _num(row[COL_ER_TAX_AMT])
        elif row[COL_ER_TAX_LABEL] in ER_UI_ETT_TAXES:
            result["er_ui_ett"] += _num(row[COL_ER_TAX_AMT])

        if row[COL_EARNINGS_LABEL] == "Total":
            result["pay"]            = _num(row[COL_PAY])
            result["employer_taxes"] = _num(row[COL_ER_TAX_AMT])
            result["deductions"]     = _num(row[COL_DEDUCTIONS_AMT])
            result["reimbursements"] = _num(row[COL_REIMB_AMT])
            result["ee_benefits"]    = _num(row[COL_EE_BENEFITS_AMT])
            result["er_benefits"]    = _num(row[COL_ER_BENEFITS_AMT])
            result["net_pay"]        = _num(row[COL_NET_PAY])
            break

    if not from_date or not to_date:
        raise ValueError(f"{xlsx_path}: could not find a 'From:'/'To:' date range")
    if from_date != to_date:
        raise ValueError(
            f"{xlsx_path}: report spans {from_date} to {to_date} — this parser only "
            f"supports a single check-date report (run one export per check date)."
        )
    result["check_date"] = normalize_check_date(from_date)

    # Deductions/reimbursements/employer benefits contributions aren't handled
    # yet (no client has needed them so far) — rather than silently drop them
    # from the journal, fail loudly so a human maps the correct QB account
    # before this run gets booked. EE benefits deductions (401(k) withholding)
    # *are* handled — see _build_journal — since Needles Studio has these.
    for key in ("deductions", "reimbursements", "er_benefits"):
        if abs(result[key]) > 0.01:
            raise ValueError(
                f"{xlsx_path}: nonzero '{key}' (${result[key]:,.2f}) — this parser doesn't "
                f"know which QB account that maps to yet. Add handling before running this payroll."
            )

    return result


def _build_journal(cfg: dict, parsed: dict, check_date: str) -> list:
    bank = cfg["payroll_bank_account"]
    rows = [
        make_row(check_date, cfg["wages_account"], debit=parsed["pay"]),
        make_row(check_date, cfg["employer_tax_account"], debit=parsed["employer_taxes"]),
        make_row(check_date, bank, credit=parsed["net_pay"], memo="Netpay"),
        make_row(check_date, bank, credit=round(parsed["ee_edd"], 2), memo="EDD"),
        make_row(check_date, bank, credit=round(parsed["ee_irs"] + parsed["er_irs"], 2), memo="IRS"),
        make_row(check_date, bank, credit=round(parsed["er_ui_ett"], 2), memo="UI/ETT"),
    ]
    # 401(k) employee withholding (Roth/Traditional) — deducted from gross pay
    # same as taxes, remitted to the plan provider the same way as the tax
    # buckets above: another bank credit, memo'd separately for QB visibility.
    if parsed["ee_benefits"] > 0.01:
        rows.append(make_row(check_date, bank, credit=round(parsed["ee_benefits"], 2), memo="Benefits"))
    return rows


def run_square_payroll(args, config_name):
    if len(args) < 1:
        print("Usage: python payroll.py square_payroll <company_totals.xlsx> --config <client.json>")
        sys.exit(1)

    xlsx_path = args[0]
    cfg = load_config(config_name)

    print(f"Client:  {cfg['client_name']}")
    print(f"XLSX:    {xlsx_path}")

    parsed = parse_workbook(xlsx_path)
    check_date = parsed["check_date"]
    print(f"Check Date:  {check_date}")

    print(f"\n--- Parsed Values ---")
    print(f"  Gross pay:       ${parsed['pay']:,.2f}")
    print(f"  Employer taxes:  ${parsed['employer_taxes']:,.2f}")
    print(f"  Net pay:         ${parsed['net_pay']:,.2f}")
    print(f"  EDD:             ${parsed['ee_edd']:,.2f}")
    print(f"  IRS:             ${parsed['ee_irs'] + parsed['er_irs']:,.2f}")
    print(f"  UI/ETT:          ${parsed['er_ui_ett']:,.2f}")
    if parsed["ee_benefits"] > 0.01:
        print(f"  Benefits (401k): ${parsed['ee_benefits']:,.2f}")

    rows = _build_journal(cfg, parsed, check_date)
    total_d, total_c = check_balance(rows)
    if abs(total_d - total_c) > 0.01:
        print(f"⚠️  JE out of balance: debits ${total_d:,.2f} vs credits ${total_c:,.2f}")

    print_journal_table(rows, cfg["client_name"], check_date)
    if _qb_confirm(cfg["client_name"]):
        append_payroll_log(cfg.get("payroll_key") or cfg["client_name"], cfg["client_name"], check_date, rows)
        append_digest_log(cfg["client_name"], check_date)
        archive_payroll_pdf(xlsx_path, cfg["client_name"], check_date)
